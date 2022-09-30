from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook


def parse_vocab(filepath):
    wb = load_workbook(Path(filepath))
    total = {}
    for sheet in wb.worksheets:
        cols = [[cell.value for cell in col] for col in sheet.columns if col[0].value]
        # sanity check
        if len(cols) % 3:
            exit(ValueError('the number of columns should be a multiple of 3'))

        triples = {}
        for i in range(0, len(cols), 3):
            current = None
            legend = None
            tr = []
            for j in range(len(cols[i])):
                a, b, c = cols[i][j], cols[i + 1][j], cols[i + 2][j]
                if j == 0:
                    legend = (a, b, c)
                    current = a
                elif a:
                    tr.append((a, b, c))
            if tr:
                triples[current] = {'legend': legend, 'words': tr}
        if triples:
            total[sheet.title] = triples

    return total


def parse_tagged_sentences(tagged_path):
    parsed = {}
    for f in tagged_path.glob('*.xlsx'):
        wb = load_workbook(f)
        for sheet in wb.worksheets:
            title = sheet.title
            if not 'sentences' in title:
                continue
            if title not in parsed:
                parsed[title] = defaultdict(int)

            rows = [[cell.value for cell in row] for row in sheet.rows]
            for i in range(0, len(rows), 4):
                words = [r for r in rows[i] if r]
                pos = [r for r in rows[i+1] if r]
                pairs = tuple(zip(words, pos))
                parsed[title][pairs] += 1
    return parsed